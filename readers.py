import openpyxl
import os
import numpy as np
from obspy import UTCDateTime, read


def Read_catalog(catalog_file, file_type):

    event_list = {
        'time': [],
        'latitude': [],
        'longitude': [],
        'depth': [],
        'magnitude': [],
        'event_id': [],
        'reloc_quality': []
    } 
    if file_type == '.dat':
        with open(catalog_file, 'r') as f:
            lines = f.readlines()
            for line in lines:
                parts = line.strip().split(',')
                event_list['time'].append(UTCDateTime(parts[0]))
                event_list['latitude'].append(float(parts[1]))
                event_list['longitude'].append(float(parts[2]))
                event_list['depth'].append(float(parts[3]))
                event_list['magnitude'].append(float(parts[4]))

    elif file_type == '.xlsx':
        # Read Excel file
        # Time, Longitude, Latitude, Depth, Magnitude, Event_ID, Relocation_Quality
        # Skip the first row (header)
        wb = openpyxl.load_workbook(catalog_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
            try:
                event_list['time'].append(UTCDateTime(row[0].value))
                event_list['longitude'].append(float(row[1].value))
                event_list['latitude'].append(float(row[2].value))
                event_list['depth'].append(float(row[3].value))
                event_list['magnitude'].append(float(row[4].value))
                event_list['event_id'].append(int(row[5].value))
                event_list['reloc_quality'].append(row[6].value)
            except TypeError:
                print(row)


    # Convert latitude, longitude, depth, and magnitude to numpy arrays
    event_list['latitude'] = np.array(event_list['latitude'])
    event_list['longitude'] = np.array(event_list['longitude'])
    event_list['depth'] = np.array(event_list['depth'])
    event_list['magnitude'] = np.array(event_list['magnitude'])

    return event_list


def Read_phase(phase_file, file_type):
    # Read phase file and extract P and S arrival times
    # A typical phase file has the following format (PAL-Output):
    # Event line: time(can be in two formats: 2023-05-04T23:23:05.759223Z or 20230504232305.91),latitude,longitude,depth,magnitude
    # Phase line: station_name,P_arrival_time,S_arrival_time, P_weight,S_weight
    # A typical phase file has the following format (TomoATT-Input):
    # Event line: Event-index year month day hour minute second latitude longitude dep_km magnitude num_recs id_event
    # Phase line: Event-index receiver-index station_name latitude longitude elevation_m phase_type travel_times

    def Is_event_line(line, file_type=file_type):
        Is_event = False

        if file_type == 'Hypoinverse-Output':
            if line.startswith('20'):
                Is_event = True
        elif file_type == 'TomoATT-Input':
            # If the second item in the line is a year (4 digits), it is an event line
            parts = line.strip().split()
            if len(parts[1]) == 4:
                Is_event = True
        elif file_type == 'Picker-Output':
            if line.startswith('20'):
                Is_event = True

        return Is_event

    

    phase_list = {
        'event_time': [],
        'event_id': [],
        'event_location': [],
        'event_magnitude': [],
        'num_recs': [],
        'phase_stations':[],
        'station_locations': [],
        'P_arrival_times': [],
        'S_arrival_times': [],
        'P_travel_times': [],
        'S_travel_times': [],
        'P_weights': [],
        'S_weights': []
    }
    
    if file_type == 'Hypoinverse-Output':
        with open(phase_file, 'r') as f:
            lines = f.readlines()
            event_time = None
            for line in lines:
                if Is_event_line(line):
                    parts = line.strip().split(',')
                    try:
                        if '-' in parts[0]:
                            event_time = UTCDateTime(parts[0])
                        else:
                            event_time = UTCDateTime.strptime(parts[0], "%Y%m%d%H%M%S.%f")
                    except Exception as e:
                        print(f"Error parsing event time: {parts[0]} with error {e}")
                        event_time = None
                        continue
                    phase_list['event_time'].append(event_time)
                    if parts[-1]:
                        phase_list['event_id'].append(parts[-1])
                    phase_list['event_magnitude'].append(float(parts[-2]))
                    phase_list['event_location'].append((float(parts[1]), float(parts[2]), float(parts[3])))
                    phase_list['phase_stations'].append([])
                    phase_list['P_arrival_times'].append([])
                    phase_list['S_arrival_times'].append([])
                    phase_list['P_weights'].append([])
                    phase_list['S_weights'].append([])
                else:
                    parts = line.strip().split(',')
                    if event_time is not None:
                        phase_list['phase_stations'][-1].append(parts[0])
                        try:
                            P_time = UTCDateTime(parts[1])
                        except Exception as e:
                            P_time = None
                        phase_list['P_arrival_times'][-1].append(P_time)
                        try:
                            S_time = UTCDateTime(parts[2])
                        except Exception as e:
                            S_time = None
                        phase_list['S_arrival_times'][-1].append(S_time)
                        try:
                            P_weight = float(parts[5])
                        except Exception as e:
                            P_weight = None
                        phase_list['P_weights'][-1].append(P_weight)
                        try:
                            S_weight = float(parts[6][:-1])
                        except Exception as e:
                            S_weight = None
                        phase_list['S_weights'][-1].append(S_weight)

    elif file_type == 'TomoATT-Input':
        with open(phase_file, 'r') as f:
            lines = f.readlines()
            event_time = None
            for line in lines:
                if Is_event_line(line):
                    parts = line.strip().split()
                    try:
                        event_time = UTCDateTime(year=int(parts[1]), month=int(parts[2]), day=int(parts[3]),
                                                 hour=int(parts[4]), minute=int(parts[5]), second=int(float(parts[6].split('.')[0])), microsecond=int((float(parts[6]) % 1) * 1e6))
                    except Exception as e:
                        print(f"Error parsing event time with error {e}")
                        event_time = None
                        continue
                    phase_list['event_time'].append(event_time)
                    phase_list['event_id'].append(parts[-1])
                    phase_list['event_magnitude'].append(float(parts[10]))
                    phase_list['num_recs'].append(int(parts[11]))
                    phase_list['event_location'].append((float(parts[7]), float(parts[8]), float(parts[9]))) # lat,lon,dep
                    phase_list['phase_stations'].append([])
                    phase_list['station_locations'].append([])
                    phase_list['P_travel_times'].append([])
                    phase_list['S_travel_times'].append([])
                else:
                    parts = line.strip().split()
                    if event_time is not None:
                        phase_list['phase_stations'][-1].append(parts[2])
                        phase_list['station_locations'][-1].append((float(parts[3]), float(parts[4]), float(parts[5]))) # lat,lon,dep
                        try:
                            if parts[6] == 'P':
                                P_travel_time = float(parts[7])
                                phase_list['P_travel_times'][-1].append(P_travel_time)
                            if len(parts) > 8:
                                if parts[8] == 'S':
                                    S_travel_time = float(parts[9])
                                    phase_list['S_travel_times'][-1].append(S_travel_time)
                        except Exception as e:
                            print(f"Error parsing travel time with error {e}")
                            continue

    elif file_type == 'Picker-Output':
        with open(phase_file, 'r') as f:
            lines = f.readlines()
            for line in lines:
                if  Is_event_line(line):
                    parts = line.strip().split(',')
                    if parts[-1]:
                        phase_list['event_id'].append(parts[-1])
                    phase_list['num_recs'].append(int(parts[1]))
                    phase_list['phase_stations'].append([])
                    phase_list['P_arrival_times'].append([])
                    phase_list['S_arrival_times'].append([])
                else:
                    parts = line.strip().split(',')
                    phase_list['phase_stations'][-1].append(parts[0])
                    try:
                        P_time = UTCDateTime(parts[2])
                    except Exception as e:
                        P_time = None
                    phase_list['P_arrival_times'][-1].append(P_time)
                    try:
                        S_time = UTCDateTime(parts[3])
                    except Exception as e:
                        S_time = None
                    phase_list['S_arrival_times'][-1].append(S_time)

    return phase_list


def Read_common_receiver_phase_file(cr_phase_file, network_code=''):

    def Is_event_line(line):
        if line.startswith('#'):
            return True

    cr_phase_list = {
        'main_event_id': [],
        'paired_event_id': [],
        'P_station_name': [],
        'S_station_name': [],
        'P_dt': [],
        'S_dt': []
    }

    with open(cr_phase_file, 'r') as f:
        lines = f.readlines()
        for line in lines:
            parts = line.strip().split()
            if Is_event_line(line):
                try:
                    cr_phase_list['main_event_id'].append(parts[1])
                    cr_phase_list['paired_event_id'].append(parts[2])
                    cr_phase_list['P_station_name'].append([])
                    cr_phase_list['S_station_name'].append([])
                    cr_phase_list['P_dt'].append([])
                    cr_phase_list['S_dt'].append([])
                except Exception as e:
                    print(f"Error parsing event line: {line} with error {e}")
                    continue
            else:
                try:
                    station_name = parts[0]
                    if network_code:
                        station_name = network_code + '.' + station_name
                    if parts[-1] == 'P':
                        cr_phase_list['P_station_name'][-1].append(station_name)
                        dt = float(parts[1]) - float(parts[2])
                        cr_phase_list['P_dt'][-1].append(dt)
                    elif parts[-1] == 'S':
                        cr_phase_list['S_station_name'][-1].append(station_name)
                        dt = float(parts[1]) - float(parts[2])
                        cr_phase_list['S_dt'][-1].append(dt)
                except Exception as e:
                    print(f"Error parsing phase line: {line} with error {e}")
                    continue

    return cr_phase_list


def Read_faults_list(file, region):
    # Read fault list from a .kml file
    # A typical fault list file has the following format:
    # <coordinates>-121.8229,37.7301 -122.0388,37.8771</coordinates>
    # Some lines contains extra 0
    # <coordinates>-121.8229,37.7301,0 -122.0388,37.8771,0</coordinates>
    def Point_is_within_target_region(lon, lat):
        lon_min, lon_max, lat_min, lat_max = region
        if lon >= lon_min and lon <= lon_max and lat >= lat_min and lat <= lat_max:
            return True
        else:
            return False
        
    
    faults_list = {
        'index': [],
        'longitude': [],
        'latitude': []
    }

    with open(file, 'r') as f:
        lines = f.readlines()
        
        index = 0
        for line in lines:
            if '<coordinates>' in line:
                fault_lon = []
                fault_lat = []
                lon_lat = line.split('<coordinates>')[1].split('</coordinates>')[0].split()
                for item in lon_lat:
                    coordinates = item.split(',')
                    try:
                        lon = coordinates[0]
                        lat = coordinates[1]

                        if Point_is_within_target_region(float(lon), float(lat)):
                            fault_lon.append(float(lon))
                            fault_lat.append(float(lat))

                    except IndexError:
                        print("Invalid coordinates: ", coordinates)
                        continue
    
                if fault_lon and fault_lat:
                    faults_list['index'].append(index)
                    faults_list['longitude'].append(fault_lon)
                    faults_list['latitude'].append(fault_lat)
                    index += 1
    return faults_list


def Read_main_faults(folder, fault_file_list):
    main_faults = {
        'index': [],
        'longitude': [],
        'latitude': [],
        'line_width': [],
        'color': []
    }

    for index, fault_file in enumerate(fault_file_list):
        fault_file_path = os.path.join(folder, fault_file)
        if os.path.exists(fault_file_path):
            with open(fault_file_path, 'r') as f:
                lines = f.readlines()
                fault_lon = []
                fault_lat = []
                for line in lines:
                    if 'Line-width' in line:
                        main_faults['line_width'].append(line.strip().split()[1])
                    elif 'Color' in line:
                        main_faults['color'].append(line.strip().split()[1])
                    else:
                        coordinates = line.strip().split()
                        try:
                            lon = coordinates[0]
                            lat = coordinates[1]
                            fault_lon.append(float(lon))
                            fault_lat.append(float(lat))
                        except IndexError:
                            print("Invalid coordinates: ", coordinates)
                            continue
                if fault_lon and fault_lat:
                    main_faults['index'].append(index)
                    main_faults['longitude'].append(fault_lon)
                    main_faults['latitude'].append(fault_lat)
        else:
            print(f"Fault file {fault_file} not found in {folder}.")
    return main_faults


def Read_station_list(station_folder, array_name, skip_header=False):
    # Read station list from a folder
    # A typical station list file has the following format:
    # index,station_name,latitude,longitude,elevation,scaling_factor

    stations_list = {
        'index': [],
        'station_name': [],
        'latitude': [],
        'longitude': [],
        'elevation': [],
        'scaling_factor': []
    }
    station_info_filename = array_name + '.txt'
    station_info_path = os.path.join(station_folder, station_info_filename)
    if not os.path.exists(station_info_path):
        print(f"Station info file {station_info_filename} not found in {station_folder}.")
        return stations_list
    
    with open(station_info_path, 'r') as f:
        lines = f.readlines()
        if skip_header:
            lines = lines[1:]
        for line in lines:
            parts = line.strip().split(',')
            try:
                stations_list['index'].append(int(parts[0]))
                stations_list['station_name'].append(parts[1])
                stations_list['latitude'].append(float(parts[2]))
                stations_list['longitude'].append(float(parts[3]))
                stations_list['elevation'].append(float(parts[4]))
                stations_list['scaling_factor'].append(float(parts[5]))
            except (IndexError, ValueError):
                    print("Invalid line: ", line)
                    continue

    # Convert the event location to numpy array
    stations_list['latitude'] = np.array(stations_list['latitude'])
    stations_list['longitude'] = np.array(stations_list['longitude'])
    stations_list['elevation'] = np.array(stations_list['elevation'])
    stations_list['scaling_factor'] = np.array(stations_list['scaling_factor'])
    return stations_list


def Read_Hypoinverse_output(filename):
    evid_station_weight = {}

    def Is_event_line(line):
        # If the line contains 'SEQUENCE NO.', it is an event line
        if 'SEQUENCE NO.' in line:
            return True
        
    
    def Is_header_line_of_station_weight(line):
        # If the line contains 'STA NET COM', it is a header line of station weight
        if 'STA NET COM' in line:
            return True
        

    def Is_terminator_line(line):
        # If the line contains '##########', it is a terminator line
        if '##########' in line:
            return True


    with open(filename, 'r') as f:
        lines = f.readlines()
        index = -1
        station_name = ''
        P_weight = 0.0
        S_weight = 0.0
        for line in lines:
            if Is_event_line(line):
                evid = str(line.strip().split()[-1])
                evid_station_weight[evid] = []
                continue

            if Is_header_line_of_station_weight(line):
                index = 0
                continue

            if Is_terminator_line(line):
                index = -1
                continue

            if index == 0:
                parts = line.strip().split()
                index = 1
                try:
                    station_name = parts[1] + '.' + parts[0]
                    P_weight = parts[12]
                    continue
                except (IndexError, ValueError):
                    print("Invalid line: ", line)
                    continue

            if index == 1:
                parts = line.strip().split()
                index = 0
                try:
                    S_weight = parts[7]
                    evid_station_weight[evid].append((station_name, P_weight, S_weight))
                    continue
                except (IndexError, ValueError):
                    print("Invalid line: ", line)
                    continue


    return evid_station_weight


def Read_mainshocks_info(filename, skip_header=False):
    mainshocks_info = {
        'index': [],
        'magnitude': [],
        'focal_mechanism': [], # (strike, dip, rake)
        'epicenter_location': [], # (latitude, longitude)
        'fracture_init_point': [], # (latitude, longitude)
        'symbol_style': [], # (style_string, color)
        'pen_style': [] # (width, color)
    }

    with open(filename, 'r') as f:
        lines = f.readlines()
        if skip_header:
            lines = lines[1:]
        for line in lines:
            parts = line.strip().split(',')
            try:
                mainshocks_info['index'].append(int(parts[0]))
                mainshocks_info['magnitude'].append(float(parts[1]))
                mainshocks_info['focal_mechanism'].append((float(parts[2]), float(parts[3]), float(parts[4])))
                mainshocks_info['epicenter_location'].append((float(parts[5]), float(parts[6])))
                mainshocks_info['fracture_init_point'].append((float(parts[7]), float(parts[8])))
                mainshocks_info['symbol_style'].append((parts[9], parts[10]))
                mainshocks_info['pen_style'].append((parts[11], parts[12]))
            except (IndexError, ValueError):
                    print("Invalid line: ", line)
                    continue

    return mainshocks_info


def Read_stream_data(st_paths):
    # read data
    print('reading stream: {}'.format(st_paths[0]))
    try:
        st  = read(st_paths[0])
        st += read(st_paths[1])
        st += read(st_paths[2])
    except: 
        print('bad data!'); return []
    return st



